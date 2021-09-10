# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import datetime

def header(first_date,last_date,today,org):
    reference=today.strftime("%Y-%m-%dT%H%M%S")
    txt="UIB+UNOB:4+"+reference+"'\n"
    txt+="UIH+TSDUPD:D:04A+1+"+reference+"'\n"
    txt+="MSD+AAR:61'\n"
    txt+="ORG+"+org+"+++"+org+"'\n"
    txt+="HDR+81+273:"+first_date+"/"+last_date+"*45:"+reference[:-2]+"+"+reference+"'\n"
    return txt


def footer(nbr,today):
    reference=reference=today.strftime("%Y-%m-%dT%H%M%S")
    txt="UIT+1+"+str(nbr)+"'\n"
    txt+="UIZ+"+reference+"+1'"
    return txt


def header_footer(org,before,after,first_date,last_date):
    nbr=0
    writer=open(after,'w')
    today=datetime.datetime.now()
    writer.write(header(first_date, last_date, today,org))
    with open(before) as reader:
            for nbr,line in enumerate(reader):
                writer.write(line)
    writer.write(footer(nbr+6   , today))
    
class xlsx_to_TSDUPD:
    def __init__(self,path):
        self.path=path
        self.wb_stops = load_workbook(self.path['stops']) 
        self.ws_stops = self.wb_stops.active
        self.wb_prd = load_workbook(self.path['prd'])
        self.ws_prd=self.wb_prd.active
        self.wb_other_names = load_workbook(self.path['other_names'])
        self.ws_other_names = self.wb_other_names.active
        self.wb_other_informations = load_workbook(self.path['other_informations'])
        self.ws_other_informations = self.wb_other_informations.active
        self.edifact=open(self.path['edifact'],'w')
        
    def create_TSDUPD(self):
        dic_relation={}
        for line in self.ws_other_informations:
            UIC1,UIC2,time,unit,rl1,rl2,ser,brand1,brand2,time2,provider1,provider2=line
            UIC1=UIC1.value
            UIC2=UIC2.value
            time=time.value
            unit=unit.value
            rl1=rl1.value
            rl2=rl2.value
            ser=ser.value
            brand1=brand1.value
            brand2=brand2.value
            time2=time2.value
            provider1=provider1.value
            provider2=provider2.value
            if UIC1 not in dic_relation:
                dic_relation[UIC1]=[]
            dic_relation[UIC1]=dic_relation[UIC1]+[(UIC2,time,unit,rl1,rl2,ser,brand1,brand2,time2,provider1,provider2)]
        
        dic_prd={}
        for line in self.ws_prd:
            UIC,service_brand1,service_brand2,connection_time,service_provider1,service_provider2=line
            UIC=UIC.value
            service_brand1=service_brand1.value
            service_brand2=service_brand2.value
            connection_time=connection_time.value
            service_provider1=service_provider1.value
            service_provider2=service_provider2.value
            if UIC not in dic_prd:
                dic_prd[UIC]=[]
            dic_prd[UIC]=dic_prd[UIC]+[(service_brand1,service_brand2,connection_time,service_provider1,service_provider2)]
        
        dic_other_names={}
        for line in self.ws_other_names:
            UIC,country,name=line
            UIC=UIC.value
            country=country.value
            name=name.value
            if UIC not in dic_other_names:
                dic_other_names[UIC]=[]
            dic_other_names[UIC]=dic_other_names[UIC]+[(country,name)]
        
        
        for line in self.ws_stops.rows:
            typo,UIC,name,short_name,lat,lon,na,begin,end,delay,country,timezone,timezone2,resa=line
            typo=typo.value
            UIC=UIC.value
            name=name.value
            short_name=short_name.value
            lat=lat.value
            lon=lon.value
            na=na.value
            begin=begin.value
            end=end.value
            delay=delay.value
            country=country.value
            timezone=timezone.value
            timezone2=timezone2.value
            resa=resa.value
            
            txt='ALS+'+typo+'+'+UIC+':'+name
            if lat not in [None,'','None'] and lon not in [None,'','None']:
                txt+='+'+lat+'+'+lon
            txt+="'\n"
            txt+='POP+273'+':'+begin+'/'+end+"'\n"
            if delay not in [None,'','None']:
                txt+='POP+87:'+delay+"'\n"
            txt+="CNY+"+country+"'\n"
            txt+="TIZ+"+timezone+':'+timezone2+"'\n"
            if short_name not in [None,'','None']:
                txt+='IFT+X02+'+short_name+"'\n"
            
            if UIC in dic_other_names:
                for country,name in dic_other_names[UIC]:
                    txt+="IFT+AGW"
                    if country not in  [None,'','None']:
                        txt+="::::"+country
                    txt+='+'+name+"'\n"
                    
            if UIC in dic_prd:
                for service_brand1,service_brand2,connection_time,service_provider1,service_provider2 in dic_prd[UIC]:
                    txt+="PRD+:::"
                    if service_brand1 not in [None,'','None']:
                        txt+=service_brand1
                    txt+=':'
                    if service_brand2 not in [None,'','None']:
                        txt+=service_brand2
                    if connection_time in [None,'','None']:
                        connection_time=''
                        print('Warning : no connection time for ',UIC)
                    else:
                        txt+='::'+connection_time
                    if service_provider1 not in [None,'','None'] :
                         txt+='+'+service_provider1
                    elif service_provider2 not in [None,'','None']:
                        txt+='+'
                    if service_provider2 not in [None,'','None']:
                        txt+='*'+service_provider2
                    txt+="'\n"          
            
            if resa not in [None,'','None']:
                txt+='RFR+X01:'+resa+"'\n"
            PRD=False
            if UIC in dic_relation:
                for UIC2,time,unit,rl1,rl2,ser,brand1,brand2,time2,provider1,provider2 in dic_relation[UIC]:
                    if UIC2 not in [None,'','None']:
                     txt+='RFR+AWN:'+UIC2+"'\n"
                    if time not in [None,'','None']:
                        txt+='MES+'+time+':'+unit+"'\n"
                    if rl1  not in [None,'','None']:
                        txt+='RLS+'+rl1+'+'+rl2+"'\n"
                    if brand1 or brand2 or time2 or provider1 or provider2 not in [None,'','None']:
                        txt+="PRD+:::"
                        PRD=True
                        if brand1 not in [None,'','None']:
                            txt+=brand1
                        txt+=':'
                        if brand2 not in [None,'','None']:
                            txt+=brand2
                        if time2 not in [None,'','None']:
                            txt+='::'+time2
                        if provider1 not in [None,'','None'] :
                             txt+='+'+provider1
                        elif provider2 not in [None,'','None']:
                            txt+='+'
                        if provider2 not in [None,'','None']:
                            txt+='*'+provider2
                        txt+="'\n"
                  
                        
                    if ser not in [None,'','None']:
                        if not PRD:
                            txt+="PRD'\n"
                        for SER in ser.split(';')[:-1]:
                            txt+="SER+"+SER+"'\n"
      
            self.edifact.write(txt)
        self.edifact.close()
              
if __name__ == "__main__":   
    path={}
    path['stops']='./XLSX/TSDUPD_STOPS.xlsx'
    path['prd']='./XLSX/TSDUPD_MCT.xlsx'
    path['other_names']='./XLSX/TSDUPD_SYNONYMS.xlsx'
    path['other_informations']='./XLSX/TSDUPD_FOOTPATH.xlsx'
    path['edifact']='./NEW_TSDUPD/temp.r'
    tr=xlsx_to_TSDUPD(path)
    
    tr.create_TSDUPD()
    org='0000'
    begin_date="2021-01-01"
    end_date="2021-12-12"
    header_footer(org,"./NEW_TSDUPD/temp.r", "./NEW_TSDUPD/TSDUPD.r",begin_date ,end_date )
    print("Finished")
