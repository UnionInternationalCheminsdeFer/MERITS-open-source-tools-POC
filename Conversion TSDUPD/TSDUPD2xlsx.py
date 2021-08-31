# -*- coding: utf-8 -*-
import glob
from openpyxl import Workbook

class TSDUPD_to_csv:
    def __init__(self,path):
        self.path=path
        self.wb_stop = Workbook()
        self.ws_stop = self.wb_stop.active
        self.wb_prd = Workbook()
        self.ws_prd = self.wb_prd.active
        self.wb_other_names = Workbook()
        self.ws_other_names = self.wb_other_names.active
        self.wb_other_informations = Workbook()
        self.ws_other_informations = self.wb_other_informations.active
        self.inserer=False
        self.cpt_id=0
        self.max_tag=[]
        self.model={'RFR+:':'',
                                 'MES+':'',
                                 'MES+:':'',
                                 'RLS+':'',
                                 'RLS++':'',
                                 'SER+':'',
                                 'PRD+:::':'',
                                 'PRD+::::':'',
                                 'PRD+::::::':'',
                                 'PRD++':'',
                                 'PRD++*':''
                                 }
                 
    def process_TSDUPD(self,fichier):
        data=open(fichier).read().split("'\n")
        for line in data:
            TAG=line[:3]
            """
            TODO
            """
            if TAG in ['UIB','UIH','MSD','ORG','HDR','TCE','UIT','UIZ']:
                continue
            elif TAG=='ALS':
                self.ALS(line)
            elif TAG=='POP':
                self.POP(line)
            elif TAG=='CNY':
                self.CNY(line)
            elif TAG=='TIZ':
                self.TIZ(line)
            elif TAG=='IFT':
                self.IFT(line)
            elif TAG=='MES':
                self.MES(line)
            elif TAG=='RLS':
                self.RLS(line)
            elif TAG=='RFR':
                self.RFR(line)
            elif TAG=='PRD':
                self.PRD(line)
            elif TAG=='SER':
                self.SER(line)
            elif line!='':
                print("Cas non traité:",line)
        self.write_sql()
        
        
    def new_stop(self):
        self.liste_other_informations=[]
        self.other_informations=self.model.copy()
        self.other_names=[]
        self.liste_prd=[]
        self.info_prd={'PRD+:::':'',
                       'PRD+::::':'',
                       'PRD+::::::':'',
                       'PRD++':'',
                       'PRD++*':''}
        self.info_stop={'ALS+':'', 'ALS++':'', 'ALS++:':'', 'IFT+X02':'','ALS+++':'', 'ALS++++':'', 
                        'ALS+++++':'','273 POP+':'', 'begin':'','end':'', '87POP+':'', '87POP+:':'','CNY+':'','TIZ+':'','TIZ+:':'','RFR+X01':''}
        
    def write_sql(self):     
        if self.other_informations!=self.model and self.other_informations not in self.liste_other_informations:
            self.liste_other_informations.append(self.other_informations)
        
        tab_stops =[]
        for X in ['ALS+', 'ALS++', 'ALS++:', 'IFT+X02','ALS+++', 'ALS++++', 'ALS+++++','begin','end','87POP+:','CNY+','TIZ+','TIZ+:','RFR+X01']:
            tab_stops.append(self.info_stop[X])
        UIC=self.info_stop['ALS++']
        self.ws_stop.append(tab_stops)
              
        if self.other_names!=[]:
            for country,name in self.other_names:
                self.ws_other_names.append([UIC,country,name])
                
           
        for row in self.liste_other_informations:
            if row !=self.model:
                tab=[UIC]
                for X in ['RFR+:','MES+','MES+:','RLS+','RLS++','SER+','PRD+:::','PRD+::::','PRD+::::::','PRD++','PRD++*']:
                    tab.append(row[X])
                self.ws_other_informations.append(tab)        
        
        for row in self.liste_prd:
            if row !={'PRD+:::':'',
                       'PRD+::::':'',
                       'PRD+::::::':'',
                       'PRD++':'',
                       'PRD++*':''}:
                tab=[UIC]
                for X in ['PRD+:::','PRD+::::','PRD+::::::','PRD++','PRD++*']:
                    tab.append(row[X])
                self.ws_prd.append(tab)
    
    def ALS(self,line):
        if self.inserer:
            self.write_sql()
        else:
            self.inserer=True
        self.new_stop()
        champs=self.process_line(line)
        for tag,valeur in champs:
            self.info_stop[tag]=valeur
        self.sql_stops=""
        self.sql_relations = ""
        self.sql_prd=''
        
    def POP(self,line):
        if line[4:7]=='273':
            pre_tag='273'
        elif line[4:6]=='87':
            pre_tag='87'
        else:
            raise Exception("Erreur de tag POP :", line)
        champs=self.process_line(line)
        for tag,valeur in champs:
            if pre_tag=='273' and tag=='POP+:':
                begin,end=valeur.split('/')
                self.info_stop['begin']=begin
                self.info_stop['end']=end
            else:
                self.info_stop[pre_tag+tag]=valeur
        
    def CNY(self,line):
        champs=self.process_line(line)
        for tag,valeur in champs:
            self.info_stop[tag]=valeur
    
    def TIZ(self,line):
        champs=self.process_line(line)
        for tag,valeur in champs:
            self.info_stop[tag]=valeur
    
    def IFT(self,line):
        if "IFT+X02+" in line:
            self.info_stop['IFT+X02']=line.split('+')[-1]
        elif "IFT+AGW" in line:
            if ":" not in line:
                country=''
                name=line.split('+')[-1]
            else:
                country,name=line.split(':')[-1].split('+')
            self.other_names.append([country,name])
            
    def MES(self,line):
        champs=self.process_line(line)
        for tag,valeur in champs:
            self.other_informations[tag]=valeur
        
    def RLS(self,line):
        champs=self.process_line(line)
        for tag,valeur in champs:
            self.other_informations[tag]=valeur
        
    def RFR(self,line):
        if "RFR+X01" in line:
            self.info_stop['RFR+X01']=line.split(':')[-1]
        else:
            if self.other_informations!=self.model:
                self.liste_other_informations.append(self.other_informations)
            self.other_informations=self.model.copy()
            champs=self.process_line(line)
            for tag,valeur in champs:
                self.other_informations[tag]=valeur
    
            
    def PRD(self,line):
        champs=self.process_line(line)
        if self.other_informations!=self.model:
            if self.other_informations['PRD+:::']!='' \
            or self.other_informations['PRD+::::']!='' \
            or self.other_informations['PRD+::::::']!='' \
            or self.other_informations['PRD++']!=''\
            or self.other_informations['PRD++*']!='':
                self.liste_other_informations.append(self.other_informations)
                self.other_informations=self.model.copy()
            for tag,valeur in champs:
                self.other_informations[tag]=valeur
        else:
            for tag,valeur in champs:
                self.info_prd[tag]=valeur
            self.liste_prd.append(self.info_prd)
            self.info_prd={'PRD+:::':'',
                           'PRD+::::':'',
                           'PRD+::::::':'',
                           'PRD++':'',
                           'PRD++*':''}
          
    def SER(self,line):
        champs=self.process_line(line)
        for tag,valeur in champs:
            self.other_informations[tag]=self.other_informations[tag]+valeur+";"
 
    def process_line(self,data):
        tag=data[:4]
        data=data[4:]
        data=data.replace('?*','²')
        data=data.replace('?+','¤')
        liste_finale=[]
        data=[i.split(':') for i in data.split('+')]   
        for Xcpt,X in enumerate(data):
           
            cle=tag+'+'*Xcpt
            for Y in X:
                if '*' in Y:
                    dat=Y.split('*')
                    for Z in dat:
                        if Z!='':
                            Z=Z.replace('²','?*')
                            Z=Z.replace('¤','?+')
                            liste_finale.append((cle,Z))
                        cle=cle+'*'
                elif Y!='':
                    Y=Y.replace('²','?*')
                    Y=Y.replace('¤','?+')
                    liste_finale.append((cle,Y))
                cle=cle+':'
        return liste_finale
        
    def create_csv_file(self):
        for file in glob.glob(self.path['edifact']+'*'):
            if "TSDUPD" in file.upper():
                print(file)
                self.process_TSDUPD(file)
              
        self.wb_stop.save(self.path['stops'])
        
        self.wb_prd.save(self.path['prd'])
        self.wb_other_names.save(self.path['other_names'])
        self.wb_other_informations.save(self.path['other_informations'])
        return True

if __name__ == "__main__":   
    path={}
    path['stops']='./XLSX/TSDUPD_STOPS.xlsx'
    path['prd']='./XLSX/TSDUPD_MCT.xlsx'
    path['other_names']='./XLSX/TSDUPD_OTHER_NAMES.xlsx'
    path['other_informations']='./XLSX/TSDUPD_FOOTPATH.xlsx'
    path['edifact']='./TSDUPD/'
    tr=TSDUPD_to_csv(path)
    tr.create_csv_file()
    print("Finished")
